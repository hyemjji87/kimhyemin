"""
LF몰 제휴 시크릿 페이지 큐레이션 대시보드
"""

import streamlit as st
import pandas as pd
import numpy as np
import json
import io
from datetime import datetime

# ─────────────────────────────────────────────
# 페이지 설정
# ─────────────────────────────────────────────
st.set_page_config(
    page_title="LF몰 시크릿 큐레이션",
    page_icon="🏷️",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─────────────────────────────────────────────
# CSS
# ─────────────────────────────────────────────
st.markdown("""
<style>
html, body, [class*="css"] {
    font-family: 'Pretendard', 'Apple SD Gothic Neo', 'Noto Sans KR', sans-serif;
}
section[data-testid="stSidebar"] { background: #1a2a40; }
section[data-testid="stSidebar"] * { color: #FFFFFF !important; }
section[data-testid="stSidebar"] p,
section[data-testid="stSidebar"] span,
section[data-testid="stSidebar"] div { color: #FFFFFF !important; }
section[data-testid="stSidebar"] label {
    color: #D0DFF0 !important; font-size: 0.82rem; font-weight: 600;
}
section[data-testid="stSidebar"] .stMarkdown { color: #FFFFFF !important; }
section[data-testid="stSidebar"] h2,
section[data-testid="stSidebar"] h3 { color: #FFFFFF !important; font-weight: 700; }
section[data-testid="stSidebar"] .stCaption { color: #B0C8E0 !important; }
section[data-testid="stSidebar"] small { color: #B0C8E0 !important; }
section[data-testid="stSidebar"] .badge-ok   { background: #1a5c30; color: #7FFFA0 !important; border-radius:4px; padding:2px 8px; font-size:.72rem; font-weight:700; }
section[data-testid="stSidebar"] .badge-wait { background: #5c4a00; color: #FFE066 !important; border-radius:4px; padding:2px 8px; font-size:.72rem; font-weight:700; }
.main .block-container { background: #F5F6F8; padding-top: 1.5rem; }

.header-banner {
    background: linear-gradient(135deg, #1E3A5F 0%, #0F1B2D 70%);
    border-radius: 12px; padding: 1.6rem 2rem; margin-bottom: 1.5rem;
    display: flex; justify-content: space-between; align-items: center;
}
.header-title  { color: #FFF; font-size: 1.4rem; font-weight: 700; letter-spacing: -0.02em; }
.header-sub    { color: #7BA3CC; font-size: 0.82rem; margin-top: .25rem; }
.header-badge  {
    background: rgba(255,255,255,.08); border: 1px solid rgba(255,255,255,.15);
    border-radius: 20px; padding: .3rem .9rem; color: #A8D4FF; font-size: .78rem;
}

.kpi-row { display: flex; gap: 12px; margin-bottom: 1.5rem; flex-wrap: wrap; }
.kpi-card {
    flex: 1; min-width: 130px;
    background: #FFF; border-radius: 10px; padding: 1rem 1.2rem;
    border-left: 3px solid #1E3A5F; box-shadow: 0 1px 4px rgba(0,0,0,.06);
}
.kpi-label { color: #6B7A8D; font-size: .7rem; font-weight: 600; letter-spacing: .06em; text-transform: uppercase; }
.kpi-value { color: #1E3A5F; font-size: 1.6rem; font-weight: 700; margin-top: .2rem; line-height: 1; }
.kpi-sub   { color: #9BA8B5; font-size: .7rem; margin-top: .25rem; }

.sec-hdr {
    background: #FFF; border-radius: 10px 10px 0 0;
    padding: .85rem 1.2rem; border-bottom: 2px solid #F0F2F5;
    display: flex; align-items: center; gap: .6rem;
}
.sec-num {
    background: #1E3A5F; color: #FFF; border-radius: 6px;
    width: 24px; height: 24px; display: inline-flex;
    align-items: center; justify-content: center;
    font-size: .72rem; font-weight: 700; flex-shrink: 0;
}
.sec-title { color: #1E3A5F; font-size: .88rem; font-weight: 700; }
.sec-desc  { color: #8A96A3; font-size: .72rem; margin-left: auto; }

.tbl-wrap {
    background: #FFF; border-radius: 0 0 10px 10px;
    padding: 1rem; box-shadow: 0 1px 4px rgba(0,0,0,.06); margin-bottom: 1.5rem;
}

.badge-ok   { background: #E6F4EA; color: #2D7D46; border-radius: 4px; padding: 2px 8px; font-size: .7rem; font-weight: 600; }
.badge-wait { background: #FFF3CD; color: #8A6914; border-radius: 4px; padding: 2px 8px; font-size: .7rem; font-weight: 600; }

.w-chip         { display:inline-block; border-radius:12px; padding:.15rem .6rem; font-size:.68rem; font-weight:600; margin-right:4px; }
.w-chip.trend   { background:#FFF0F6; color:#D63E8A; }
.w-chip.vol     { background:#F0FFF4; color:#2D7D46; }
.w-chip.growth  { background:#FFF8E1; color:#B45309; }

.info-box { background:#EEF4FF; border-radius:8px; padding:.8rem 1rem; color:#2255A4; font-size:.78rem; margin:.5rem 0; }
.warn-box { background:#FFF8E1; border-radius:8px; padding:.8rem 1rem; color:#7B5A00; font-size:.78rem; margin:.5rem 0; }

.stTabs [data-baseweb="tab-list"] {
    gap:4px; background:#F0F2F5; border-radius:8px; padding:4px;
}
.stTabs [data-baseweb="tab"] {
    border-radius:6px; font-size:.82rem; font-weight:600; color:#6B7A8D; padding:6px 14px;
}
.stTabs [aria-selected="true"] { background:#FFF !important; color:#1E3A5F !important; }

.stDownloadButton > button {
    background:#1E3A5F !important; color:white !important;
    border-radius:8px !important; border:none !important;
    font-weight:600 !important; font-size:.82rem !important;
}
.stDownloadButton > button:hover { background:#2A5080 !important; }
</style>
""", unsafe_allow_html=True)


# ─────────────────────────────────────────────
# 유틸
# ─────────────────────────────────────────────
def _to_float(v):
    """콤마 포함 문자열, None, NaN 모두 float으로 안전 변환"""
    try:
        if pd.isna(v): return None
    except Exception:
        pass
    try:
        return float(str(v).replace(',', ''))
    except Exception:
        return None

def fmt_amt(v):
    v = _to_float(v)
    if v is None or v == 0: return "0"
    if abs(v) >= 1e8: return f"{v/1e8:.1f}억"
    if abs(v) >= 1e4: return f"{v/1e4:.0f}만"
    return f"{v:,.0f}"

def fmt_num(v):
    v = _to_float(v)
    if v is None: return "-"
    return f"{int(v):,}"

def fmt_pct(v):
    v = _to_float(v)
    if v is None: return "-"
    sign = "+" if v > 0 else ""
    return f"{sign}{v:.1f}%"

def mall_week_to_code(week_str):
    """'2025년 06월 3주차' → '25_6_3'"""
    try:
        s = week_str.replace("년","_").replace("월","_").replace("주차","").strip()
        parts = [p.strip() for p in s.split("_") if p.strip()]
        y = parts[0][-2:]
        m = str(int(parts[1]))
        w = parts[2]
        return f"{y}_{m}_{w}"
    except:
        return week_str

def code_to_mall_week(code, year=2025):
    """'25_6_3' → '2025년 06월 3주차'"""
    try:
        parts = code.split("_")
        y = 2000 + int(parts[0])
        m = int(parts[1])
        w = parts[2]
        return f"{y}년 {m:02d}월 {w}주차"
    except:
        return code

def get_prev_week_code(code):
    """'26_6_3' → '25_6_3' (전년 동주차)"""
    try:
        parts = code.split("_")
        return f"{int(parts[0])-1}_{parts[1]}_{parts[2]}"
    except:
        return None

def build_week_map(df_pivot):
    """피벗 E열↔G열: {날짜str: 주차코드}"""
    wmap = {}
    for _, row in df_pivot.iterrows():
        dval = row.iloc[4]
        wcode = row.iloc[6]
        if pd.notna(dval) and pd.notna(wcode):
            try:
                d = pd.to_datetime(dval)
                wmap[d.strftime('%Y-%m-%d')] = str(wcode).strip()
            except:
                pass
    return wmap

def build_affiliate_map(df_pivot):
    """피벗 B열↔C열: {코드: 제휴사명}"""
    amap = {}
    for _, row in df_pivot.iterrows():
        code = str(row.iloc[1]).strip() if pd.notna(row.iloc[1]) else ''
        name = str(row.iloc[2]).strip() if pd.notna(row.iloc[2]) else ''
        if code and name and code not in ('<제휴사 구분>', 'nan'):
            amap[code] = name
    return amap

def get_vat_col(df):
    """거래액_VAT제외 컬럼 반환. 없으면 거래액/1.1 계산 후 반환"""
    if '거래액_VAT제외' in df.columns:
        return pd.to_numeric(df['거래액_VAT제외'], errors='coerce')
    return pd.to_numeric(df['거래액'], errors='coerce') / 1.1

def tag_week(df, week_map):
    """정산일시일 → 주차코드 컬럼 추가"""
    df = df.copy()
    df['_date'] = pd.to_datetime(df['정산일시일']).dt.strftime('%Y-%m-%d')
    df['_week'] = df['_date'].map(week_map)
    return df

def df_to_excel_bytes(df, sheet_name='Sheet1'):
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        ws = writer.sheets[sheet_name]
        hdr_fill = PatternFill(fill_type='solid', fgColor='1E3A5F')
        hdr_font = Font(bold=True, color='FFFFFF')
        hdr_align = Alignment(horizontal='center', vertical='center')
        for cell in ws[1]:
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = hdr_align
        for col_cells in ws.columns:
            max_w = max(len(str(c.value)) if c.value else 0 for c in col_cells)
            ws.column_dimensions[col_cells[0].column_letter].width = min(max_w + 2, 28)
    buf.seek(0)
    return buf.getvalue()

def dfs_to_json_bytes(data_dict):
    out = {}
    for key, df in data_dict.items():
        if df is not None:
            dc = df.copy()
            for col in dc.select_dtypes(include=['datetime64']).columns:
                dc[col] = dc[col].astype(str)
            out[key] = dc.to_dict(orient='records')
        else:
            out[key] = []
    return json.dumps(out, ensure_ascii=False, indent=2).encode('utf-8')


# ─────────────────────────────────────────────
# 로딩 함수 (캐시)
# ─────────────────────────────────────────────
@st.cache_data(show_spinner=False)
def load_mall_csv(file_bytes):
    try:
        df = pd.read_csv(io.BytesIO(file_bytes), encoding='utf-16', sep='\t')
        df.columns = [c.strip() for c in df.columns]
        if '거래액' in df.columns:
            df['거래액'] = pd.to_numeric(
                df['거래액'].astype(str).str.replace(',', ''), errors='coerce'
            )
        # 주차코드 컬럼 추가 (피벗 없이 자체 변환)
        if '결제_주차' in df.columns:
            df['_week_code'] = df['결제_주차'].apply(mall_week_to_code)
        return df, None
    except Exception as e:
        return None, str(e)

@st.cache_data(show_spinner=False)
def load_affiliate_excel(file_bytes, sheet_hint='26'):
    """
    인증거래액 시트 + 피벗 시트 로드.
    sheet_hint: '26' → '인증거래액_26', '25' → '인증거래액_4-6월' 등 자동 탐색
    """
    try:
        buf = io.BytesIO(file_bytes)
        import openpyxl
        wb = openpyxl.load_workbook(buf, read_only=True)
        sheets = wb.sheetnames
        wb.close()

        # 인증거래액 시트 탐색
        raw_sheet = None
        for s in sheets:
            if '인증거래액' in s:
                raw_sheet = s
                break
        if raw_sheet is None:
            return None, None, f"'인증거래액' 시트를 찾을 수 없습니다. 시트목록: {sheets}"

        buf = io.BytesIO(file_bytes)
        df_raw = pd.read_excel(buf, sheet_name=raw_sheet)

        # 피벗 시트
        df_piv = None
        if '피벗' in sheets:
            buf = io.BytesIO(file_bytes)
            df_piv = pd.read_excel(buf, sheet_name='피벗', header=None)

        return df_raw, df_piv, None
    except Exception as e:
        return None, None, str(e)


# ─────────────────────────────────────────────
# 분석 함수
# ─────────────────────────────────────────────

def analyze_mall_top20(df_mall, target_mall_week):
    """
    [분석1] 전년 몰전체 특정 주차 판매 TOP 20
    출력: 카테고리/브랜드/고객수/거래액
    (자사입점·정상이월은 몰전체 CSV에 없으므로 '-' 처리)
    """
    if df_mall is None or df_mall.empty:
        return pd.DataFrame()

    df = df_mall[df_mall['결제_주차'] == target_mall_week].copy()
    if df.empty:
        return pd.DataFrame()

    grp = df.groupby(['대카테고리명', 'ADMIN브랜드명'], as_index=False).agg(
        고객수=('주문고객수', 'sum'),
        거래액=('거래액', 'sum')
    ).sort_values('거래액', ascending=False).head(20).reset_index(drop=True)
    grp.insert(0, '순위', range(1, len(grp)+1))
    grp['자사/입점'] = '-'
    grp['정상/이월'] = '-'
    grp = grp[['순위','대카테고리명','ADMIN브랜드명','자사/입점','정상/이월','고객수','거래액']]
    grp.columns = ['순위','카테고리','브랜드','자사/입점','정상/이월','고객수','거래액']
    return grp


def analyze_affiliate_top20_vol(df_raw, week_map, target_week_code):
    """
    [분석2] 제휴 실적 분석주차 판매 TOP 20 (당월인증 Y/N 모두 포함)
    출력: 카테고리/브랜드/정상이월/고객수/거래액/객단가
    """
    if df_raw is None or df_raw.empty or not week_map:
        return pd.DataFrame()

    df = tag_week(df_raw[df_raw['정산구분'] == '판매'], week_map)
    df = df[df['_week'] == target_week_code].copy()
    if df.empty:
        return pd.DataFrame()

    df['amt'] = get_vat_col(df)
    grp = df.groupby(['물리대카테', 'Admin브랜드명', '정상이월구분'], as_index=False).agg(
        고객수=('고객번호', 'nunique'),
        거래액=('amt', 'sum')
    )
    grp['객단가'] = (grp['거래액'] / grp['고객수']).round(0)
    grp = grp.sort_values('거래액', ascending=False).head(20).reset_index(drop=True)
    grp.insert(0, '순위', range(1, len(grp)+1))
    grp.columns = ['순위','카테고리','브랜드','정상/이월','고객수','거래액','객단가']
    return grp


def analyze_affiliate_top20_growth(df_curr, df_prev, week_map_curr, week_map_prev,
                                    target_week_code):
    """
    [분석3] 전년 동주차 대비 신장률 TOP 20
    df_curr: 2026 raw / df_prev: 2025 raw
    전년 동주차: 26_6_3 → 25_6_3
    """
    if df_curr is None or df_curr.empty:
        return pd.DataFrame()

    prev_week_code = get_prev_week_code(target_week_code)

    # 당년 집계
    df_c = tag_week(df_curr[df_curr['정산구분'] == '판매'], week_map_curr)
    df_c = df_c[df_c['_week'] == target_week_code].copy()
    df_c['amt'] = get_vat_col(df_c)

    grp_c = df_c.groupby(['물리대카테', 'Admin브랜드명'], as_index=False).agg(
        고객수_당=('고객번호', 'nunique'),
        거래액_당=('amt', 'sum')
    )

    # 전년 집계 (2025 raw 있을 때)
    grp_p = pd.DataFrame(columns=['물리대카테', 'Admin브랜드명', '거래액_전'])
    if df_prev is not None and not df_prev.empty and week_map_prev and prev_week_code:
        df_p = tag_week(df_prev[df_prev['정산구분'] == '판매'], week_map_prev)
        df_p = df_p[df_p['_week'] == prev_week_code].copy()
        df_p['amt'] = get_vat_col(df_p)
        if not df_p.empty:
            grp_p = df_p.groupby(['물리대카테', 'Admin브랜드명'], as_index=False).agg(
                거래액_전=('amt', 'sum')
            )

    merged = grp_c.merge(grp_p, on=['물리대카테', 'Admin브랜드명'], how='left')
    merged['전년비(%)'] = np.where(
        merged['거래액_전'].fillna(0) > 0,
        (merged['거래액_당'] - merged['거래액_전']) / merged['거래액_전'] * 100,
        np.nan
    )
    # 전년 실적 있는 브랜드만, 최소 거래액 30만
    merged = merged[(merged['거래액_당'] >= 300000) & merged['전년비(%)'].notna()]
    merged = merged.sort_values('전년비(%)', ascending=False).head(20).reset_index(drop=True)
    merged.insert(0, '순위', range(1, len(merged)+1))
    merged = merged[['순위','물리대카테','Admin브랜드명','고객수_당','거래액_당','전년비(%)']]
    merged.columns = ['순위','카테고리','브랜드','고객수','거래액','전년비(%)']
    return merged


def score_and_select(df1, df2, df3, top_n=30, w1=0.35, w2=0.35, w3=0.30):
    """
    [분석4] 가중치 종합점수 브랜드 선정
    각 분석의 순위 역수 기반 점수화 후 가중 합산
    """
    scores = {}

    def add(df, weight, score_key):
        if df is None or df.empty:
            return
        n = len(df)
        for i, (_, row) in enumerate(df.iterrows()):
            brand = row.get('브랜드', '')
            cat   = row.get('카테고리', '')
            score = (n - i) / n * weight * 100
            if brand not in scores:
                scores[brand] = {
                    '브랜드': brand, '카테고리': cat,
                    '트렌드점수': 0.0, '볼륨점수': 0.0, '신장률점수': 0.0
                }
            scores[brand][score_key] += score

    add(df1, w1, '트렌드점수')
    add(df2, w2, '볼륨점수')
    add(df3, w3, '신장률점수')

    if not scores:
        return pd.DataFrame()

    result = pd.DataFrame(scores.values())
    result['종합점수'] = (result['트렌드점수'] + result['볼륨점수'] + result['신장률점수']).round(1)
    result['트렌드점수'] = result['트렌드점수'].round(1)
    result['볼륨점수']   = result['볼륨점수'].round(1)
    result['신장률점수'] = result['신장률점수'].round(1)
    result = result.sort_values('종합점수', ascending=False).head(top_n).reset_index(drop=True)
    result.insert(0, '순위', range(1, len(result)+1))
    return result[['순위','카테고리','브랜드','종합점수','트렌드점수','볼륨점수','신장률점수']]


def get_top_products(df_raw, week_map, target_week_code, selected_brands, top_n=10):
    """
    [분석5] 선택 브랜드의 상품코드 TOP 10 (고객수 기준)
    출력: 브랜드/자사입점/정상이월/상품코드/상품명/MD명/고객수/거래액
    """
    if df_raw is None or not selected_brands:
        return pd.DataFrame()

    df = tag_week(df_raw[df_raw['정산구분'] == '판매'], week_map)
    df = df[(df['_week'] == target_week_code) & df['Admin브랜드명'].isin(selected_brands)].copy()
    if df.empty:
        return pd.DataFrame()

    df['amt'] = get_vat_col(df)

    # 자사/입점 컬럼 처리 (제휴처구분1로 대체)
    입점col = '자사/입점' if '자사/입점' in df.columns else '제휴처구분1'

    grp_cols = ['Admin브랜드명', 입점col, '정상이월구분', '상품코드', '상품명', 'MD명']
    grp = df.groupby(grp_cols, as_index=False).agg(
        고객수=('고객번호', 'nunique'),
        거래액=('amt', 'sum')
    )

    parts = []
    for brand in selected_brands:
        b = grp[grp['Admin브랜드명'] == brand].nlargest(top_n, '고객수')
        parts.append(b)

    if not parts:
        return pd.DataFrame()

    final = pd.concat(parts, ignore_index=True)
    final = final.rename(columns={'Admin브랜드명': '브랜드', 입점col: '자사/입점', '정상이월구분': '정상/이월'})
    return final[['브랜드','자사/입점','정상/이월','상품코드','상품명','MD명','고객수','거래액']]


# ─────────────────────────────────────────────
# 세션 상태 초기화
# ─────────────────────────────────────────────
defaults = {
    'df_mall': None, 'df_aff_25': None, 'df_aff_26': None,
    'df_piv_25': None, 'df_piv_26': None,
    'week_map_25': {}, 'week_map_26': {},
    'analysis_done': False,
    'df1': pd.DataFrame(), 'df2': pd.DataFrame(),
    'df3': pd.DataFrame(), 'df4': pd.DataFrame(),
    'sel_week': '', 'sel_prev_mall_week': '',
}
for k, v in defaults.items():
    if k not in st.session_state:
        st.session_state[k] = v


# ─────────────────────────────────────────────
# 사이드바
# ─────────────────────────────────────────────
with st.sidebar:
    st.markdown("## 🏷️ 시크릿 큐레이션")
    st.markdown("---")
    st.markdown("### 📂 데이터 업로드")

    # ── ① 몰전체 CSV ──
    st.markdown("**① 2025년 몰 주차별 실적 (CSV)**")
    f_mall = st.file_uploader("mall", type=['csv'], label_visibility='collapsed', key='up_mall')
    if f_mall:
        with st.spinner("로딩..."):
            df, err = load_mall_csv(f_mall.read())
        if err:
            st.error(err)
        else:
            st.session_state.df_mall = df
            st.markdown('<span class="badge-ok">✓ 로드 완료</span>', unsafe_allow_html=True)
            st.caption(f"{len(df):,}행 · 브랜드 {df['ADMIN브랜드명'].nunique():,}개")
    else:
        st.markdown('<span class="badge-wait">⏳ 대기 중</span>', unsafe_allow_html=True)

    st.markdown("")

    # ── ② 2025 제휴 raw ──
    st.markdown("**② 2025년 분기 제휴 실적 (xlsx)**")
    f25 = st.file_uploader("aff25", type=['xlsx'], label_visibility='collapsed', key='up_aff25')
    if f25:
        with st.spinner("로딩..."):
            dr, dp, err = load_affiliate_excel(f25.read(), '25')
        if err:
            st.error(err)
        else:
            st.session_state.df_aff_25 = dr
            if dp is not None:
                st.session_state.df_piv_25 = dp
                st.session_state.week_map_25 = build_week_map(dp)
            st.markdown('<span class="badge-ok">✓ 로드 완료</span>', unsafe_allow_html=True)
            st.caption(f"{len(dr):,}행")
    else:
        st.markdown('<span class="badge-wait">⏳ 대기 중</span>', unsafe_allow_html=True)

    st.markdown("")

    # ── ③ 2026 제휴 raw ──
    st.markdown("**③ 2026년 당월 제휴 실적 (xlsx)**")
    f26 = st.file_uploader("aff26", type=['xlsx'], label_visibility='collapsed', key='up_aff26')
    if f26:
        with st.spinner("로딩..."):
            dr, dp, err = load_affiliate_excel(f26.read(), '26')
        if err:
            st.error(err)
        else:
            st.session_state.df_aff_26 = dr
            if dp is not None:
                st.session_state.df_piv_26 = dp
                st.session_state.week_map_26 = build_week_map(dp)
            st.markdown('<span class="badge-ok">✓ 로드 완료</span>', unsafe_allow_html=True)
            st.caption(f"{len(dr):,}행 · {dr['정산일시일'].min()} ~ {dr['정산일시일'].max()}")
    else:
        st.markdown('<span class="badge-wait">⏳ 대기 중</span>', unsafe_allow_html=True)

    st.markdown("---")
    st.markdown("### 📅 분석 기간 설정")

    # 2026 주차코드 목록
    avail_26 = []
    if st.session_state.df_aff_26 is not None and st.session_state.week_map_26:
        wm = st.session_state.week_map_26
        dates = pd.to_datetime(st.session_state.df_aff_26['정산일시일']).dt.strftime('%Y-%m-%d').unique()
        avail_26 = sorted(set(wm.get(d,'') for d in dates if wm.get(d,'')))

    if avail_26:
        sel_week = st.selectbox("분석 주차 (2026)", avail_26, index=len(avail_26)-1)
    else:
        sel_week = st.text_input("분석 주차 (직접 입력, 예: 26_6_3)", value='26_6_3')

    # 전년 비교 주차: 몰전체 CSV 주차 목록에서 선택
    prev_code = get_prev_week_code(sel_week)  # e.g. 25_6_3
    auto_prev_mall = code_to_mall_week(prev_code) if prev_code else ''

    avail_mall_weeks = []
    if st.session_state.df_mall is not None:
        avail_mall_weeks = sorted(st.session_state.df_mall['결제_주차'].unique())

    if avail_mall_weeks:
        default_idx = avail_mall_weeks.index(auto_prev_mall) if auto_prev_mall in avail_mall_weeks else 0
        sel_prev_mall = st.selectbox(
            "전년 비교 주차 (몰전체 기준)",
            avail_mall_weeks, index=default_idx
        )
    else:
        sel_prev_mall = st.text_input("전년 비교 주차 (예: 2025년 06월 3주차)", value=auto_prev_mall)

    st.caption(f"분석: `{sel_week}` ↔ 전년: `{mall_week_to_code(sel_prev_mall) if sel_prev_mall else '-'}`")

    st.markdown("---")
    with st.expander("⚙️ 가중치 설정"):
        w1 = st.slider("전년 트렌드", 0.0, 1.0, 0.35, 0.05)
        w2 = st.slider("제휴 볼륨",   0.0, 1.0, 0.35, 0.05)
        w3 = st.slider("신장률",       0.0, 1.0, 0.30, 0.05)
        tw = w1 + w2 + w3
        if abs(tw - 1.0) > 0.02:
            st.warning(f"합계 {tw:.2f} (합계 1.0 권장)")

    st.markdown("---")
    st.markdown("### 💾 Raw 데이터 내보내기")
    has_data = any(st.session_state[k] is not None for k in ['df_mall','df_aff_25','df_aff_26'])
    if has_data:
        json_bytes = dfs_to_json_bytes({
            'mall_2025': st.session_state.df_mall,
            'affiliate_2025': st.session_state.df_aff_25,
            'affiliate_2026': st.session_state.df_aff_26,
        })
        st.download_button(
            "📥 JSON으로 다운로드", json_bytes,
            file_name=f"lf_secret_raw_{datetime.now().strftime('%y%m%d')}.json",
            mime='application/json', use_container_width=True
        )
    else:
        st.caption("파일 업로드 후 활성화됩니다.")

    st.markdown("---")
    run_btn = st.button("🔍 분석 실행", use_container_width=True, type="primary")


# ─────────────────────────────────────────────
# 메인 영역
# ─────────────────────────────────────────────
st.markdown(f"""
<div class="header-banner">
  <div>
    <div class="header-title">🏷️ LF몰 시크릿 페이지 큐레이션 대시보드</div>
    <div class="header-sub">전년 트렌드 × 제휴 실적 × 신장률 → 게시 브랜드 & 상품 선정</div>
  </div>
  <div class="header-badge">내부 전용 · CONFIDENTIAL</div>
</div>
""", unsafe_allow_html=True)

# ── 분석 실행 ──
if run_btn:
    missing = []
    if st.session_state.df_mall is None:      missing.append("① 몰전체 CSV")
    if st.session_state.df_aff_26 is None:    missing.append("③ 2026 제휴 xlsx")
    if not st.session_state.week_map_26:      missing.append("피벗 시트 (2026 xlsx 내)")

    if missing:
        st.warning(f"필수 파일 누락: {', '.join(missing)}")
    else:
        with st.spinner("분석 중..."):
            df1 = analyze_mall_top20(st.session_state.df_mall, sel_prev_mall)
            df2 = analyze_affiliate_top20_vol(
                st.session_state.df_aff_26, st.session_state.week_map_26, sel_week
            )
            df3 = analyze_affiliate_top20_growth(
                st.session_state.df_aff_26, st.session_state.df_aff_25,
                st.session_state.week_map_26, st.session_state.week_map_25,
                sel_week
            )
            df4 = score_and_select(df1, df2, df3, top_n=30, w1=w1, w2=w2, w3=w3)

        st.session_state.df1 = df1
        st.session_state.df2 = df2
        st.session_state.df3 = df3
        st.session_state.df4 = df4
        st.session_state.sel_week = sel_week
        st.session_state.sel_prev_mall_week = sel_prev_mall
        st.session_state.analysis_done = True

        # 검증 로그
        checks = []
        checks.append(f"분석1 브랜드: {len(df1)}개 (전년 몰전체 {sel_prev_mall})")
        checks.append(f"분석2 브랜드: {len(df2)}개 (제휴 볼륨 {sel_week})")
        checks.append(f"분석3 브랜드: {len(df3)}개 (신장률 전년비교)")
        checks.append(f"최종 선정: {len(df4)}개")
        st.success("분석 완료 ✓\n" + " · ".join(checks))

# ── KPI 요약 ──
if st.session_state.analysis_done:
    w = st.session_state.sel_week
    pw = st.session_state.sel_prev_mall_week
    n1,n2,n3,n4 = (len(st.session_state[f'df{i}']) for i in range(1,5))
    st.markdown(f"""
    <div class="kpi-row">
      <div class="kpi-card">
        <div class="kpi-label">전년 트렌드 TOP</div>
        <div class="kpi-value">{n1}</div>
        <div class="kpi-sub">몰전체 · {mall_week_to_code(pw)}</div>
      </div>
      <div class="kpi-card">
        <div class="kpi-label">제휴 볼륨 TOP</div>
        <div class="kpi-value">{n2}</div>
        <div class="kpi-sub">제휴 · {w}</div>
      </div>
      <div class="kpi-card">
        <div class="kpi-label">신장률 TOP</div>
        <div class="kpi-value">{n3}</div>
        <div class="kpi-sub">전년 동주차 대비</div>
      </div>
      <div class="kpi-card" style="border-left-color:#D63E8A;">
        <div class="kpi-label">최종 선정</div>
        <div class="kpi-value" style="color:#D63E8A;">{n4}</div>
        <div class="kpi-sub">가중치 종합</div>
      </div>
    </div>
    """, unsafe_allow_html=True)

# ─────────────────────────────────────────────
# 탭
# ─────────────────────────────────────────────
tab1, tab2, tab3, tab4, tab5 = st.tabs([
    "📊 ① 전년 트렌드 TOP 20",
    "💰 ② 제휴 볼륨 TOP 20",
    "📈 ③ 신장률 TOP 20",
    "🏷️ ④ 최종 선정 30개",
    "🛍️ ⑤ 상품코드 출력",
])

def render_table_tab(tab, num, title, desc, chip_cls, chip_txt, df_key,
                     fmt_cols=None, dl_sheet='Sheet1', dl_prefix=''):
    with tab:
        st.markdown(f"""
        <div class="sec-hdr">
          <span class="sec-num">{num}</span>
          <span class="sec-title">{title}</span>
          <span class="sec-desc">
            <span class="w-chip {chip_cls}">{chip_txt}</span>{desc}
          </span>
        </div>
        <div class="tbl-wrap">
        """, unsafe_allow_html=True)

        df = st.session_state[df_key]
        if st.session_state.analysis_done and not df.empty:
            df_show = df.copy()
            if fmt_cols:
                for col, fn in fmt_cols.items():
                    if col in df_show.columns:
                        df_show[col] = df_show[col].apply(fn)
            st.dataframe(df_show, use_container_width=True, hide_index=True, height=550)
            st.download_button(
                "📥 엑셀 다운로드",
                data=df_to_excel_bytes(df, dl_sheet),
                file_name=f"{dl_prefix}_{datetime.now().strftime('%y%m%d')}.xlsx",
                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        elif st.session_state.analysis_done:
            st.markdown('<div class="warn-box">⚠️ 해당 주차 데이터 없음. 주차 선택 및 파일 확인 필요.</div>',
                        unsafe_allow_html=True)
        else:
            st.markdown('<div class="info-box">📌 사이드바에서 파일 업로드 후 <strong>분석 실행</strong>을 눌러주세요.</div>',
                        unsafe_allow_html=True)
        st.markdown("</div>", unsafe_allow_html=True)

render_table_tab(
    tab1, "1", "전년 몰전체 해당 주차 판매 TOP 20",
    "기준: 2025년 몰전체 CSV · 거래액 내림차순",
    "trend", "트렌드 35%", "df1",
    fmt_cols={'거래액': fmt_amt, '고객수': fmt_num},
    dl_sheet='전년트렌드TOP20', dl_prefix='전년트렌드TOP20'
)

render_table_tab(
    tab2, "2", "제휴 실적 분석주차 판매 TOP 20",
    "기준: 당월인증 Y/N 포함 · 판매건 · VAT 제외 · 거래액 내림차순",
    "vol", "볼륨 35%", "df2",
    fmt_cols={'거래액': fmt_amt, '객단가': fmt_amt, '고객수': fmt_num},
    dl_sheet='제휴볼륨TOP20', dl_prefix='제휴볼륨TOP20'
)

render_table_tab(
    tab3, "3", "전년 동주차 대비 신장률 TOP 20",
    "기준: 최소 거래액 30만↑ · 전년 실적 있는 브랜드만",
    "growth", "신장률 30%", "df3",
    fmt_cols={'거래액': fmt_amt, '고객수': fmt_num, '전년비(%)': fmt_pct},
    dl_sheet='신장률TOP20', dl_prefix='신장률TOP20'
)

render_table_tab(
    tab4, "4", "최종 게시 브랜드 선정 (30개)",
    "가중치 종합점수 기준 · 사이드바 ⚙️에서 가중치 조정 가능",
    "trend", "종합", "df4",
    dl_sheet='최종선정30', dl_prefix='시크릿_선정브랜드30'
)

# ── 탭5: 상품코드 ──
with tab5:
    st.markdown("""
    <div class="sec-hdr">
      <span class="sec-num">5</span>
      <span class="sec-title">브랜드별 상품코드 TOP 10 출력</span>
      <span class="sec-desc">최종 선정 브랜드 중 상품코드 추출 대상을 선택하세요</span>
    </div>
    <div class="tbl-wrap">
    """, unsafe_allow_html=True)

    if st.session_state.analysis_done and not st.session_state.df4.empty:
        brand_opts = st.session_state.df4['브랜드'].tolist()
        selected = st.multiselect(
            "상품코드 출력할 브랜드 선택",
            options=brand_opts,
            default=brand_opts[:5],
            placeholder="브랜드 선택...",
        )

        if selected:
            with st.spinner("상품 데이터 추출 중..."):
                df_prod = get_top_products(
                    st.session_state.df_aff_26,
                    st.session_state.week_map_26,
                    st.session_state.sel_week,
                    selected, top_n=10
                )

            if not df_prod.empty:
                df_show = df_prod.copy()
                df_show['거래액'] = df_show['거래액'].apply(fmt_amt)
                df_show['고객수'] = df_show['고객수'].apply(fmt_num)
                st.dataframe(df_show, use_container_width=True, hide_index=True, height=520)
                col_dl, col_info = st.columns([1, 3])
                with col_dl:
                    st.download_button(
                        "📥 상품코드 엑셀 다운로드",
                        data=df_to_excel_bytes(df_prod, '상품코드_TOP10'),
                        file_name=f"시크릿_상품코드_{datetime.now().strftime('%y%m%d')}.xlsx",
                        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                    )
                with col_info:
                    st.caption(f"선택 {len(selected)}개 브랜드 · 브랜드별 고객수 기준 TOP 10")
            else:
                st.markdown('<div class="warn-box">⚠️ 선택 브랜드의 해당 주차 상품 데이터 없음.</div>',
                            unsafe_allow_html=True)
        else:
            st.markdown('<div class="info-box">⬆️ 브랜드를 선택하면 상품코드가 출력됩니다.</div>',
                        unsafe_allow_html=True)
    else:
        st.markdown('<div class="info-box">📌 ④ 최종 선정 분석 완료 후 활성화됩니다.</div>',
                    unsafe_allow_html=True)

    st.markdown("</div>", unsafe_allow_html=True)

# ── 푸터 ──
st.markdown("---")
st.markdown(
    f"<div style='text-align:center;color:#9BA8B5;font-size:.72rem;'>"
    f"LF몰 제휴팀 내부 전용 · 무단 배포 금지 · "
    f"Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}</div>",
    unsafe_allow_html=True
)
