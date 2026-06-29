import streamlit as st
import pandas as pd
import numpy as np
import json
import io
from datetime import datetime
import warnings
warnings.filterwarnings('ignore')

st.set_page_config(
    page_title="LF몰 인증페이지 전시 추천",
    layout="wide",
    initial_sidebar_state="expanded"
)

st.markdown("""
<style>
  .main { background:#f8f9fa; }
  .metric-card {
    background:white; border-radius:10px; padding:16px 20px;
    box-shadow:0 1px 4px rgba(0,0,0,0.08); margin-bottom:8px;
  }
  .metric-label { font-size:12px; color:#6b7280; font-weight:500; margin-bottom:4px; }
  .metric-value { font-size:22px; font-weight:700; color:#1e3a5f; }
  .metric-sub   { font-size:12px; color:#9ca3af; margin-top:2px; }
  .section-title {
    font-size:16px; font-weight:700; color:#1e3a5f;
    border-left:4px solid #1e3a5f; padding-left:10px; margin:18px 0 10px;
  }
  .badge-yellow { background:#fff3cd; color:#856404; padding:2px 8px; border-radius:4px; font-size:11px; }
  .badge-green  { background:#d4edda; color:#155724; padding:2px 8px; border-radius:4px; font-size:11px; }
  .badge-star   { background:#fff9c4; color:#7d6608; padding:2px 8px; border-radius:4px; font-size:11px; }
  .top5-card {
    background:white; border-radius:8px; padding:12px 16px;
    box-shadow:0 1px 3px rgba(0,0,0,0.1); margin-bottom:8px;
    border-left:4px solid #1e3a5f;
  }
  .glossary-box {
    background:#f0f4f8; border-radius:8px; padding:16px 20px;
    font-size:13px; color:#374151; margin-top:24px;
    border:1px solid #d1dbe8;
  }
  .glossary-box b { color:#1e3a5f; }
</style>
""", unsafe_allow_html=True)

GLOSSARY = """
<div class="glossary-box">
<b>📖 용어 설명</b><br><br>
<b>기준선(전체 NB율)</b> : 제휴 전체 거래액 중 신규+WIN-BACK 고객이 차지하는 비중. 편차의 기준점.<br>
<b>NB(New Buyer)</b> : 신규 + WIN-BACK 고객을 합친 개념. 인증페이지 전시의 핵심 타겟.<br>
<b>브랜드내 NB비중(%)</b> : 해당 브랜드 거래액 중 신규+WIN-BACK 고객이 차지하는 비중.<br>
<b>편차(%p)</b> : 브랜드 NB비중 − 전체 기준선 NB율. 높을수록 신규/윈백 고객이 평균보다 많이 구매한 브랜드.<br>
<b>NB거래액(원)</b> : 해당 브랜드에서 신규+WIN-BACK 고객이 구매한 총 거래액 (VAT 제외).<br>
<b>NB고객수(명)</b> : 해당 브랜드를 구매한 신규+WIN-BACK 고객의 수 (중복 제거).<br>
<b>NB객단가(원)</b> : NB거래액 ÷ NB고객수. 1인당 평균 구매금액.<br>
<b>NB전체대비 비중(%)</b> : 전체 NB거래액 중 해당 브랜드가 차지하는 비중. 구좌가 하나일 때 볼륨 기준으로 참고.<br>
<b>제휴비중(%)</b> : 제휴 전체 거래액 중 해당 브랜드 비중.<br>
<b>몰비중(%)</b> : 몰전체 거래액 중 해당 브랜드 비중.<br>
<b>몰vs제휴 편차(%p)</b> : 제휴비중 − 몰비중. (+)면 몰 평균보다 제휴 채널에서 더 잘 팔리는 브랜드.<br>
<b>⚠ 객단가이슈</b> : NB고객수 3명 이하. 소수 고객의 고단가 구매일 수 있어 재현성이 낮음.<br>
<b>★ 몰튀는</b> : 몰vs제휴 편차 ±0.5%p 이상 & 제휴비중 0.3% 이상. 제휴 채널에 특화된 브랜드.
</div>
"""

def parse_pivot(df_raw):
    pivot = df_raw.dropna(how='all')
    date_map, code_map = {}, {}
    for _, row in pivot.iterrows():
        if pd.notna(row.iloc[4]) and pd.notna(row.iloc[6]):
            try:
                dt = pd.to_datetime(row.iloc[4])
                date_map[dt.strftime('%Y-%m-%d')] = str(row.iloc[6])
            except:
                pass
        if pd.notna(row.iloc[8]) and pd.notna(row.iloc[9]):
            code_map[str(row.iloc[8]).strip()] = str(row.iloc[9]).strip()
    return date_map, code_map

def build_json_from_raw(df_raw_all: dict) -> dict:
    df = df_raw_all['인증거래액_26'].copy()
    df_cert = df_raw_all['인증회원_26'].copy()
    df_mall = df_raw_all['몰전체_26'].copy()
    df_pivot_raw = df_raw_all['피벗']

    date_map, code_map = parse_pivot(df_pivot_raw)

    df['정산일시일'] = df['정산일시일'].astype(str)
    df['주차코드'] = df['정산일시일'].map(date_map)

    latest_date = df['정산일시일'].max()
    latest_wk = date_map.get(latest_date, '')
    all_weeks = sorted(df['주차코드'].dropna().unique().tolist())

    # SSNGCD03 항상 제외
    df = df[df['제휴처구분4'] != 'SSNGCD03']

    df_sale = df[(df['정산구분'] == '판매') & (df['당월인증'] == 'Y')]
    df_nb = df_sale[df_sale['기존/win-back/신규'].isin(['신규', 'WIN-BACK'])]

    total_rev = df_sale['거래액_VAT제외'].sum()
    nb_rev_total = df_nb['거래액_VAT제외'].sum()
    baseline_nb_rate = round(nb_rev_total / total_rev * 100, 2) if total_rev > 0 else 0

    brd_total = df_sale.groupby('Admin브랜드명').agg(
        rev_total=('거래액_VAT제외', 'sum'),
        cat=('물리대카테', lambda x: x.mode()[0] if len(x) > 0 else ''),
    ).reset_index()

    brd_nb = df_nb.groupby('Admin브랜드명').agg(
        nb_rev=('거래액_VAT제외', 'sum'),
        nb_buyers=('고객번호', 'nunique'),
    ).reset_index()

    brd = brd_total.merge(brd_nb, on='Admin브랜드명', how='left')
    brd['nb_rev'] = brd['nb_rev'].fillna(0)
    brd['nb_buyers'] = brd['nb_buyers'].fillna(0).astype(int)
    brd['nb_rate'] = (brd['nb_rev'] / brd['rev_total'] * 100).round(2)
    brd['diff'] = (brd['nb_rate'] - baseline_nb_rate).round(2)
    brd['nb_arpu'] = np.where(brd['nb_buyers'] > 0, (brd['nb_rev'] / brd['nb_buyers']).round(0), 0)
    brd['nb_pct_total'] = (brd['nb_rev'] / nb_rev_total * 100).round(2) if nb_rev_total > 0 else 0
    brd['af_pct'] = (brd['rev_total'] / total_rev * 100).round(2)

    # 몰전체 비중
    df_mall_work = df_mall.copy()
    hdr_idx = None
    for i, row in df_mall_work.iterrows():
        if any('브랜드' in str(v) for v in row.values if pd.notna(v)):
            hdr_idx = i
            break
    if hdr_idx is not None:
        df_mall_work.columns = df_mall_work.iloc[hdr_idx].tolist()
        df_mall_work = df_mall_work.iloc[hdr_idx+1:].copy()
    else:
        df_mall_work.columns = df_mall_work.iloc[0].tolist()
        df_mall_work = df_mall_work.iloc[1:].copy()

    brd_col = next((c for c in df_mall_work.columns if '브랜드' in str(c)), None)
    rev_col = next((c for c in df_mall_work.columns if '거래액' in str(c)), None)
    if brd_col and rev_col:
        df_mall_work[rev_col] = pd.to_numeric(df_mall_work[rev_col], errors='coerce').fillna(0)
        mall_brd = df_mall_work.groupby(brd_col)[rev_col].sum().reset_index()
        mall_brd.columns = ['Admin브랜드명', '거래액']
        mall_total = mall_brd['거래액'].sum()
        mall_brd['mall_pct'] = (mall_brd['거래액'] / mall_total * 100).round(2) if mall_total > 0 else 0
    else:
        mall_brd = pd.DataFrame(columns=['Admin브랜드명', 'mall_pct'])

    brd = brd.merge(mall_brd[['Admin브랜드명', 'mall_pct']], on='Admin브랜드명', how='left')
    brd['mall_pct'] = brd['mall_pct'].fillna(0)
    brd['mall_diff'] = (brd['af_pct'] - brd['mall_pct']).round(2)

    def make_note(row):
        notes = []
        if row['nb_buyers'] <= 3 and row['nb_rev'] > 0:
            notes.append('⚠객단가이슈')
        if abs(row['mall_diff']) >= 0.5 and row['af_pct'] >= 0.3:
            notes.append('★몰튀는')
        return ' '.join(notes)

    brd['비고'] = brd.apply(make_note, axis=1)
    brd_filtered = brd[brd['nb_rev'] >= 300000].sort_values('diff', ascending=False).reset_index(drop=True)

    # SKU: 편차 10%p 이상 모든 브랜드 대상으로 미리 계산
    sku_candidate_brands = brd_filtered[brd_filtered['diff'] >= 10]['Admin브랜드명'].tolist()
    sku_data = {}
    for brand in sku_candidate_brands:
        df_brd_nb = df_nb[df_nb['Admin브랜드명'] == brand]
        sku = df_brd_nb.groupby(['상품코드', '상품명']).agg(
            nb_buyers=('고객번호', 'nunique'),
            nb_rev=('거래액_VAT제외', 'sum'),
        ).reset_index().sort_values('nb_buyers', ascending=False).head(10)
        sku_data[brand] = sku.to_dict('records')

    brd_rows = []
    for _, r in brd_filtered.iterrows():
        brd_rows.append({
            'brand': r['Admin브랜드명'],
            'cat': r['cat'],
            'nb_rev': round(r['nb_rev']),
            'nb_buyers': int(r['nb_buyers']),
            'nb_arpu': round(r['nb_arpu']),
            'nb_rate': round(r['nb_rate'], 2),
            'diff': round(r['diff'], 2),
            'nb_pct_total': round(r['nb_pct_total'], 2),
            'af_pct': round(r['af_pct'], 2),
            'mall_pct': round(r['mall_pct'], 2),
            'mall_diff': round(r['mall_diff'], 2),
            'note': r['비고'],
        })

    # 주차별 데이터 (분석 주차 선택용)
    week_data = {}
    for wk in all_weeks:
        df_wk = df[(df['주차코드'] == wk) & (df['정산구분'] == '판매') & (df['당월인증'] == 'Y')]
        df_wk_nb = df_wk[df_wk['기존/win-back/신규'].isin(['신규', 'WIN-BACK'])]
        wk_total = df_wk['거래액_VAT제외'].sum()
        wk_nb = df_wk_nb['거래액_VAT제외'].sum()
        if wk_total > 0:
            wk_baseline = round(wk_nb / wk_total * 100, 2)
            wk_brd_total = df_wk.groupby('Admin브랜드명').agg(rev_total=('거래액_VAT제외','sum'), cat=('물리대카테', lambda x: x.mode()[0] if len(x)>0 else '')).reset_index()
            wk_brd_nb = df_wk_nb.groupby('Admin브랜드명').agg(nb_rev=('거래액_VAT제외','sum'), nb_buyers=('고객번호','nunique')).reset_index()
            wk_brd = wk_brd_total.merge(wk_brd_nb, on='Admin브랜드명', how='left')
            wk_brd['nb_rev'] = wk_brd['nb_rev'].fillna(0)
            wk_brd['nb_buyers'] = wk_brd['nb_buyers'].fillna(0).astype(int)
            wk_brd['nb_rate'] = (wk_brd['nb_rev'] / wk_brd['rev_total'] * 100).round(2)
            wk_brd['diff'] = (wk_brd['nb_rate'] - wk_baseline).round(2)
            wk_brd['nb_arpu'] = np.where(wk_brd['nb_buyers']>0, (wk_brd['nb_rev']/wk_brd['nb_buyers']).round(0), 0)
            wk_brd['nb_pct_total'] = (wk_brd['nb_rev'] / wk_nb * 100).round(2) if wk_nb > 0 else 0
            wk_brd['af_pct'] = (wk_brd['rev_total'] / wk_total * 100).round(2)
            wk_brd = wk_brd.merge(mall_brd[['Admin브랜드명','mall_pct']], on='Admin브랜드명', how='left')
            wk_brd['mall_pct'] = wk_brd['mall_pct'].fillna(0)
            wk_brd['mall_diff'] = (wk_brd['af_pct'] - wk_brd['mall_pct']).round(2)
            wk_brd['비고'] = wk_brd.apply(make_note, axis=1)
            wk_filtered = wk_brd[wk_brd['nb_rev'] >= 300000].sort_values('diff', ascending=False)

            wk_sku = {}
            for brand in wk_filtered[wk_filtered['diff'] >= 10]['Admin브랜드명'].tolist():
                df_b = df_wk_nb[df_wk_nb['Admin브랜드명'] == brand]
                s = df_b.groupby(['상품코드','상품명']).agg(nb_buyers=('고객번호','nunique'), nb_rev=('거래액_VAT제외','sum')).reset_index().sort_values('nb_buyers', ascending=False).head(10)
                wk_sku[brand] = s.to_dict('records')

            week_data[wk] = {
                'baseline_nb_rate': wk_baseline,
                'total_nb_rev': round(wk_nb),
                'total_nb_buyers': int(df_wk_nb['고객번호'].nunique()),
                'brd_rows': [{'brand':r['Admin브랜드명'],'cat':r['cat'],'nb_rev':round(r['nb_rev']),'nb_buyers':int(r['nb_buyers']),'nb_arpu':round(r['nb_arpu']),'nb_rate':round(r['nb_rate'],2),'diff':round(r['diff'],2),'nb_pct_total':round(r['nb_pct_total'],2),'af_pct':round(r['af_pct'],2),'mall_pct':round(r['mall_pct'],2),'mall_diff':round(r['mall_diff'],2),'note':r['비고']} for _,r in wk_filtered.iterrows()],
                'sku_top10': wk_sku,
            }

    result = {
        'meta': {
            'generated_at': datetime.now().strftime('%Y-%m-%d %H:%M'),
            'data_range': f"{df['정산일시일'].min()} ~ {latest_date}",
            'latest_week': latest_wk,
            'all_weeks': all_weeks,
            'baseline_nb_rate': baseline_nb_rate,
            'total_rev': round(total_rev),
            'nb_rev_total': round(nb_rev_total),
            'brand_count': len(brd_filtered),
        },
        'cert_page': {
            'baseline_nb_rate': baseline_nb_rate,
            'total_nb_rev': round(nb_rev_total),
            'total_nb_buyers': int(df_nb['고객번호'].nunique()),
            'brd_rows': brd_rows,
            'sku_top10': sku_data,
        },
        'week_data': week_data,
    }
    return result

def build_xlsx(brd_rows, baseline, period_label):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = f"인증페이지_전시추천_{period_label}"

    ws.merge_cells('A1:M1')
    ws['A1'] = f"인증페이지 전시 추천 — {period_label} | 신규+WIN-BACK 브랜드 편차 분석"
    ws['A1'].font = Font(bold=True, size=13, color='FFFFFF')
    ws['A1'].fill = PatternFill('solid', start_color='1E3A5F')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 26

    ws.merge_cells('A2:M2')
    ws['A2'] = f"기준선(전체 NB율): {baseline:.1f}% | 당월인증=Y, 판매, NB거래액 30만원 이상 | 편차 내림차순"
    ws['A2'].font = Font(size=10, color='374151')
    ws['A2'].alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[2].height = 18

    headers = ['순위','브랜드','카테고리','브랜드내\nNB비중(%)','편차\n(%p)','NB거래액\n(원)','NB고객수\n(명)','NB객단가\n(원)','NB전체대비\n비중(%)','제휴비중\n(%)','몰비중\n(%)','몰vs제휴\n편차(%p)','비고']
    hdr_fill = PatternFill('solid', start_color='1E3A5F')
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=ci, value=h)
        c.font = Font(bold=True, color='FFFFFF', size=10)
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[3].height = 30

    fill_yellow = PatternFill('solid', start_color='FFF3CD')
    fill_star   = PatternFill('solid', start_color='FFF9C4')
    fill_green  = PatternFill('solid', start_color='D4EDDA')
    fill_sum    = PatternFill('solid', start_color='1E3A5F')
    thin = Side(style='thin', color='DEE2E6')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    total_nb_rev = sum(r['nb_rev'] for r in brd_rows)
    total_nb_buyers = sum(r['nb_buyers'] for r in brd_rows)

    for ri, row in enumerate(brd_rows, 4):
        vals = [ri-3, row['brand'], row['cat'], row['nb_rate'], row['diff'],
                row['nb_rev'], row['nb_buyers'], row['nb_arpu'],
                row['nb_pct_total'], row['af_pct'], row['mall_pct'], row['mall_diff'], row['note']]
        row_fill = fill_yellow if '⚠객단가이슈' in row['note'] else (fill_star if '★몰튀는' in row['note'] else None)
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.border = border
            c.alignment = Alignment(horizontal='center', vertical='center')
            if row_fill:
                c.fill = row_fill
            if ci in (6, 8):
                c.number_format = '#,##0'
            elif ci in (4, 5, 9, 10, 11, 12):
                c.number_format = '0.00'
            if ci == 5 and isinstance(v, (int, float)) and v >= 10:
                c.fill = fill_green

    sr = len(brd_rows) + 4
    sum_vals = ['합계','전체 브랜드','',f'{baseline:.2f}','',total_nb_rev,total_nb_buyers,'',100.0,'','','','']
    for ci, v in enumerate(sum_vals, 1):
        c = ws.cell(row=sr, column=ci, value=v)
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = fill_sum
        c.border = border
        c.alignment = Alignment(horizontal='center', vertical='center')
        if ci == 6:
            c.number_format = '#,##0'
        elif ci in (4, 9):
            c.number_format = '0.00'

    legend_row = sr + 2
    for i, (txt, f) in enumerate([('⚠ 고객수 3명 이하 — 객단가 이슈, 재현성 낮음', fill_yellow),
                                   ('★ 제휴비중이 몰전체 대비 ±0.5%p 이상 튀는 브랜드', fill_star),
                                   ('편차 +10%p 이상 — 신규WB 집중 브랜드', fill_green)]):
        cr = legend_row + i
        ws.cell(row=cr, column=1).fill = f
        ws.cell(row=cr, column=1).border = border
        c = ws.cell(row=cr, column=2, value=txt)
        c.font = Font(size=10)
        ws.merge_cells(start_row=cr, start_column=2, end_row=cr, end_column=6)

    widths = [6,24,13,11,10,17,11,14,13,11,11,12,16]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()

def build_sku_xlsx(selected_brands, sku_data, brd_rows_dict):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)
    hdr_fill = PatternFill('solid', start_color='1E3A5F')

    for ri, brand in enumerate(selected_brands, 1):
        brow = brd_rows_dict.get(brand, {})
        ws2 = wb.create_sheet(title=f"#{ri}_{brand[:15]}")
        ws2.merge_cells('A1:F1')
        ws2['A1'] = f"#{ri} {brand} | NB고객수 {brow.get('nb_buyers',0)}명 | 편차 {brow.get('diff',0):+.2f}%p"
        ws2['A1'].font = Font(bold=True, color='FFFFFF', size=12)
        ws2['A1'].fill = hdr_fill
        ws2['A1'].alignment = Alignment(horizontal='center')

        for ci, h in enumerate(['순위','상품코드','상품명','NB고객수(명)','NB거래액(원)'], 1):
            c = ws2.cell(row=2, column=ci, value=h)
            c.font = Font(bold=True, color='FFFFFF')
            c.fill = hdr_fill
            c.alignment = Alignment(horizontal='center')

        for si, sku in enumerate(sku_data.get(brand, []), 1):
            ws2.cell(row=si+2, column=1, value=si)
            ws2.cell(row=si+2, column=2, value=sku.get('상품코드',''))
            ws2.cell(row=si+2, column=3, value=sku.get('상품명',''))
            ws2.cell(row=si+2, column=4, value=sku.get('nb_buyers',0))
            c5 = ws2.cell(row=si+2, column=5, value=round(sku.get('nb_rev',0)))
            c5.number_format = '#,##0'

        ws2.column_dimensions['A'].width = 6
        ws2.column_dimensions['B'].width = 18
        ws2.column_dimensions['C'].width = 40
        ws2.column_dimensions['D'].width = 14
        ws2.column_dimensions['E'].width = 16

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()

def main():
    st.title("🏷️ LF몰 인증페이지 전시 추천 시스템")

    with st.sidebar:
        st.markdown("### 📁 데이터 업로드")
        raw_file = st.file_uploader("raw Excel 업로드", type=['xlsx'],
                                    help="시트: 인증거래액_26 / 인증회원_26 / 몰전체_26 / 피벗")
        json_file = st.file_uploader("또는 기존 JSON 업로드", type=['json'])
        st.divider()
        st.markdown("### 🗓️ 분석 기간")
        period_mode = st.radio("기간 선택", ["MTD 전체", "주차별"], index=0)
        selected_week = None
        st.divider()
        st.markdown("### 🎯 전시 추천 필터")
        diff_threshold = st.slider("편차 기준 (%p 이상)", 0, 30, 10, step=1)
        min_buyers = st.number_input("최소 NB 고객수 (명 이상)", min_value=1, value=1, step=1)

    data = None

    if raw_file:
        with st.spinner("Raw 데이터 처리 중..."):
            try:
                sheets = ['인증거래액_26','인증회원_26','몰전체_26','피벗']
                raw_all = {}
                for s in sheets:
                    raw_all[s] = pd.read_excel(raw_file, sheet_name=s, engine='openpyxl',
                                               header=0 if s != '몰전체_26' else None)
                data = build_json_from_raw(raw_all)
                st.sidebar.success(f"✅ 생성 완료 | {data['meta']['data_range']}")
                json_bytes = json.dumps(data, ensure_ascii=False, indent=2).encode('utf-8')
                end_date = data['meta']['data_range'].split('~')[-1].strip().replace('-','')
                st.sidebar.download_button("💾 JSON 다운로드", data=json_bytes,
                                           file_name=f"dashboard_data_{end_date}.json",
                                           mime='application/json')
            except Exception as e:
                st.sidebar.error(f"오류: {e}")
                st.stop()

    elif json_file:
        try:
            data = json.load(json_file)
            st.sidebar.success("✅ JSON 로드 완료")
        except Exception as e:
            st.sidebar.error(f"JSON 오류: {e}")
            st.stop()

    if data is None:
        st.info("👈 왼쪽에서 raw Excel 또는 JSON 파일을 업로드하세요.")
        return

    # 주차 선택 드롭박스 (데이터 로드 후)
    all_weeks = data.get('meta', {}).get('all_weeks', [])
    if period_mode == "주차별" and all_weeks:
        with st.sidebar:
            selected_week = st.selectbox("주차 선택", all_weeks,
                                         index=len(all_weeks)-1)

    # 현재 사용할 데이터 결정
    if period_mode == "주차별" and selected_week and 'week_data' in data:
        cp = data['week_data'].get(selected_week, data.get('cert_page', {}))
        period_label = selected_week.replace('_', '-')
    else:
        cp = data.get('cert_page', {})
        period_label = data.get('meta', {}).get('data_range', 'MTD')

    meta = data.get('meta', {})
    brd_rows = cp.get('brd_rows', [])
    sku_data = cp.get('sku_top10', {})
    baseline = cp.get('baseline_nb_rate', 0)

    tab1, tab2 = st.tabs(["📊 전시 추천 분석", "🔑 브랜드별 SKU"])

    # ══ TAB 1 ══════════════════════════════════════════════════════════════════
    with tab1:
        st.markdown(f'<div class="section-title">분석 기준: {period_label} | 기준선(전체 NB율): <b>{baseline:.2f}%</b></div>',
                    unsafe_allow_html=True)

        total_nb_rev = cp.get('total_nb_rev', 0)
        total_nb_buyers = cp.get('total_nb_buyers', 0)
        df_all = pd.DataFrame(brd_rows)
        reco_count = len(df_all[(df_all['diff'] >= diff_threshold) & (df_all['nb_buyers'] >= min_buyers)]) if len(df_all) > 0 else 0

        c1, c2, c3, c4 = st.columns(4)
        with c1:
            st.markdown(f'<div class="metric-card"><div class="metric-label">NB 총 거래액</div><div class="metric-value">{total_nb_rev:,.0f}원</div><div class="metric-sub">신규+WIN-BACK | VAT 제외</div></div>', unsafe_allow_html=True)
        with c2:
            st.markdown(f'<div class="metric-card"><div class="metric-label">NB 고객수</div><div class="metric-value">{total_nb_buyers:,}명</div><div class="metric-sub">unique 구매고객</div></div>', unsafe_allow_html=True)
        with c3:
            st.markdown(f'<div class="metric-card"><div class="metric-label">분석 브랜드 수</div><div class="metric-value">{len(brd_rows)}개</div><div class="metric-sub">NB 거래액 30만 이상</div></div>', unsafe_allow_html=True)
        with c4:
            st.markdown(f'<div class="metric-card"><div class="metric-label">추천 후보 브랜드</div><div class="metric-value">{reco_count}개</div><div class="metric-sub">편차 ≥{diff_threshold}%p & 고객수 ≥{min_buyers}명</div></div>', unsafe_allow_html=True)

        st.divider()

        if len(df_all) == 0:
            st.warning("데이터가 없습니다.")
            return

        df_filtered = df_all[(df_all['diff'] >= diff_threshold) & (df_all['nb_buyers'] >= min_buyers)]\
            .sort_values('nb_buyers', ascending=False).reset_index(drop=True)
        df_filtered.insert(0, '순위', range(1, len(df_filtered)+1))

        col_map = {'brand':'브랜드','cat':'카테고리','nb_rate':'NB비중(%)','diff':'편차(%p)',
                   'nb_rev':'NB거래액(원)','nb_buyers':'NB고객수(명)','nb_arpu':'NB객단가(원)',
                   'nb_pct_total':'NB전체대비(%)','af_pct':'제휴비중(%)','mall_pct':'몰비중(%)',
                   'mall_diff':'몰vs제휴(%p)','note':'비고'}
        df_show = df_filtered.rename(columns=col_map)

        def style_table(df):
            styles = pd.DataFrame('', index=df.index, columns=df.columns)
            for i, row in df.iterrows():
                note = str(row.get('비고',''))
                if '⚠객단가이슈' in note:
                    styles.loc[i, :] = 'background-color: #FFF3CD'
                elif '★몰튀는' in note:
                    styles.loc[i, :] = 'background-color: #FFF9C4'
                if row.get('편차(%p)', 0) >= 10:
                    styles.loc[i, '편차(%p)'] = 'background-color: #D4EDDA; font-weight: bold'
            return styles

        disp_cols = ['순위','브랜드','카테고리','NB비중(%)','편차(%p)','NB거래액(원)',
                     'NB고객수(명)','NB객단가(원)','NB전체대비(%)','제휴비중(%)','몰비중(%)','몰vs제휴(%p)','비고']
        disp_cols = [c for c in disp_cols if c in df_show.columns]

        st.markdown(f'<div class="section-title">전시 추천 브랜드 — {len(df_filtered)}개</div>', unsafe_allow_html=True)

        styled = df_show[disp_cols].style.apply(style_table, axis=None).format({
            'NB거래액(원)':'{:,.0f}','NB객단가(원)':'{:,.0f}',
            'NB비중(%)':'{:.2f}','편차(%p)':'{:.2f}','NB전체대비(%)':'{:.2f}',
            '제휴비중(%)':'{:.2f}','몰비중(%)':'{:.2f}','몰vs제휴(%p)':'{:.2f}',
        })
        st.dataframe(styled, height=500, width=1400)

        st.markdown("""
        <div style="display:flex; gap:12px; margin-top:8px; font-size:12px;">
          <span class="badge-yellow">⚠ 객단가이슈: NB고객수 ≤3명</span>
          <span class="badge-star">★ 몰튀는: 제휴비중 몰대비 ±0.5%p 이상</span>
          <span class="badge-green">● 편차 +10%p 이상</span>
        </div>
        """, unsafe_allow_html=True)

        # Excel 다운로드
        if len(df_filtered) > 0:
            st.divider()
            xlsx_bytes = build_xlsx(
                df_all[(df_all['diff'] >= diff_threshold) & (df_all['nb_buyers'] >= min_buyers)]
                .sort_values(['diff','nb_buyers'], ascending=[False,False]).to_dict('records'),
                baseline, period_label.replace(' ','_').replace('~','-')[:20]
            )
            st.download_button("📥 전시 추천 Excel 다운로드", data=xlsx_bytes,
                               file_name=f"{period_label}_인증페이지_전시추천.xlsx",
                               mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        # 용어 설명
        st.markdown(GLOSSARY, unsafe_allow_html=True)

    # ══ TAB 2 ══════════════════════════════════════════════════════════════════
    with tab2:
        st.markdown('<div class="section-title">브랜드 선택 → SKU TOP 10</div>', unsafe_allow_html=True)

        # 편차 기준 이상 브랜드 목록
        df_candidate = df_all[(df_all['diff'] >= diff_threshold) & (df_all['nb_buyers'] >= min_buyers)]\
            .sort_values('nb_buyers', ascending=False)

        if len(df_candidate) == 0:
            st.warning("조건에 맞는 브랜드가 없습니다. 왼쪽 필터를 조정해 보세요.")
        else:
            brand_options = df_candidate['brand'].tolist()
            selected_brands = st.multiselect(
                "브랜드 선택 (복수 선택 가능)",
                options=brand_options,
                default=brand_options[:5] if len(brand_options) >= 5 else brand_options,
                help="편차 기준 이상 브랜드 중 선택하세요"
            )

            if not selected_brands:
                st.info("위에서 브랜드를 선택하세요.")
            else:
                brd_rows_dict = {r['brand']: r for r in brd_rows}

                for ri, brand in enumerate(selected_brands, 1):
                    row = brd_rows_dict.get(brand, {})
                    note_badge = ""
                    if '⚠객단가이슈' in str(row.get('note','')):
                        note_badge = ' <span class="badge-yellow">⚠ 객단가이슈</span>'
                    if '★몰튀는' in str(row.get('note','')):
                        note_badge += ' <span class="badge-star">★ 몰튀는</span>'

                    st.markdown(f"""
                    <div class="top5-card">
                      <b>#{ri} {brand}</b>&nbsp;|&nbsp;{row.get('cat','')}
                      {note_badge}
                      <br><small style="color:#6b7280;">
                        NB고객수 <b>{row.get('nb_buyers',0)}명</b> &nbsp;|&nbsp;
                        편차 <b style="color:#155724;">{row.get('diff',0):+.2f}%p</b> &nbsp;|&nbsp;
                        NB거래액 {row.get('nb_rev',0):,.0f}원 &nbsp;|&nbsp;
                        NB객단가 {row.get('nb_arpu',0):,.0f}원
                      </small>
                    </div>
                    """, unsafe_allow_html=True)

                    sku_list = sku_data.get(brand, [])
                    if sku_list:
                        df_sku = pd.DataFrame(sku_list)
                        df_sku.insert(0, '순위', range(1, len(df_sku)+1))
                        df_sku = df_sku.rename(columns={'nb_buyers':'NB고객수(명)','nb_rev':'NB거래액(원)','상품코드':'상품코드','상품명':'상품명'})
                        df_sku['NB거래액(원)'] = df_sku['NB거래액(원)'].apply(lambda x: round(x))
                        st.dataframe(df_sku, hide_index=True, width=1400)
                    else:
                        st.info(f"'{brand}' SKU 데이터가 없습니다. Raw Excel을 업로드하면 자동 계산됩니다.")
                    st.divider()

                # SKU Excel 다운로드
                if any(sku_data.get(b) for b in selected_brands):
                    brd_rows_dict = {r['brand']: r for r in brd_rows}
                    sku_xlsx = build_sku_xlsx(selected_brands, sku_data, brd_rows_dict)
                    st.download_button(
                        f"📥 선택 브랜드 SKU Excel 다운로드 ({len(selected_brands)}개)",
                        data=sku_xlsx,
                        file_name=f"{period_label}_선택브랜드_SKU.xlsx",
                        mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                    )

                # 용어 설명
                st.markdown(GLOSSARY, unsafe_allow_html=True)

if __name__ == '__main__':
    main()
